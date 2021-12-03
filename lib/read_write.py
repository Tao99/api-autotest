#!/usr/bin/env python3.9
# encoding: utf-8

import json
import openpyxl
import yaml
from yamlinclude import YamlIncludeConstructor

from config import setting

# 用于yaml文件嵌套
YamlIncludeConstructor.add_to_loader_class(loader_class=yaml.FullLoader, base_dir=setting.TESTCASE_DIE)


class Files:
    def __init__(self, file: str):
        self.path = file

    def read_yaml(self):
        """
        读取yaml文件
        :return: dict
        """
        with open(self.path, encoding='utf-8') as f:
            result = yaml.load(stream=f, Loader=yaml.FullLoader)
            return result

    def write_yaml(self, data):
        """
        写入yaml文件
        :param data: [{'database': 'Hs-api', 'port': 21020}, {'url': 'www.baidu.com'}]
        """
        with open(self.path, encoding='utf-8', mode='w') as f:
            yaml.dump(data, stream=f, allow_unicode=True)

    def get_data(self):
        data = self.read_yaml()
        return data

    def read_json(self):
        """
        读取json文件
        :return: dict
        """
        with open(self.path, 'r', encoding='utf-8') as fp:
            data = json.load(fp)
        return data

    def write_json(self, data):
        """
        写入json文件
        :param data: [{'age': 18, 'name': 'lily'}, {'age': 28, 'name': 'lucy'}]
        """
        with open(self.path, 'w', encoding='utf-8') as fp:
            json.dump(data, fp, ensure_ascii=False, indent=2)

    def read_excel(self):
        """
        读取Excel文件
        :return: 包含多个列表的列表，例如 [[],[],[]……]
        """
        table = openpyxl.load_workbook(self.path)
        sheet_name = table.sheetnames
        sheet = table[sheet_name[0]]  # 默认读取第一个sheet页
        rows = sheet.rows
        sheet_list = []
        for row in rows:
            row_list = []
            for cell in row:
                row_list.append(cell.value)
            sheet_list.append(row_list)
        return sheet_list

    def write_excel(self, data: list):
        """
        按行写入excel
        :param data: [["name", "age"], ["lily", 18], ["lucy", 20]]
        """
        excel = openpyxl.Workbook()
        sheet = excel.create_sheet('sheet1', 0)  # 创建sheet，并指定为第一个sheet
        for row_no, row in enumerate(data):
            for col_no, value in enumerate(row):
                sheet.cell(row_no + 1, col_no + 1).value = value
        excel.save(self.path)


if __name__ == '__main__':
    file1 = "E:/file/file.json"
    file2 = "E:/file/config.yml"
    file3 = "E:/file/file2.xlsx"
    data1 = [{'age': 18, 'name': 'lily'}, {'age': 28, 'name': 'lucy'}]
    data2 = [{'database': 'Hs-api', 'port': 21020}, {'url': 'www.baidu.com'}]
    data3 = [["name", "age"], ["lily", 18], ["lucy", 20]]
    Files(file1).write_json(data1)
    Files(file2).write_yaml(data2)
    Files(file3).write_excel(data3)
    print(Files(file1).read_json())
    print(Files(file2).read_yaml())
    print(Files(file3).read_excel())
